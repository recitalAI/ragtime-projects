import json
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
import os 

def json_to_xlsx(path : str):
    # Extract the directory path and filename from the JSON path
    json_dirname = os.path.dirname(path)
    json_filename = os.path.basename(path)
    # Remove the ".json" extension
    xlsx_filename = os.path.splitext(json_filename)[0] + ".xlsx"
    # Construct the full path for the XLSX file in the same directory
    xlsx_path = os.path.join(json_dirname, xlsx_filename)
    # Open the JSON file
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Find the maximum number of chunks for a question
    max_chunks = max(len(item['chunks']['items']) for item in data['items'])

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write column headers
    worksheet['A1'] = 'Question'
    column = 2
    for i in range(max_chunks):
        worksheet.cell(row=1, column=column, value=f'Chunk {i+1}')
        column += 1
    worksheet.cell(row=1, column=column, value='LLM answers')
    column += 1
    worksheet.cell(row=1, column=column, value='Human evaluations')

    # Function to adjust column width
    def set_column_width(worksheet, column, max_length=100):
        column_letter = get_column_letter(column)
        max_width = max(
            len(str(cell.value)) for cell in worksheet[column_letter]
        ) if worksheet[column_letter] else 0
        if max_width > 0:
            worksheet.column_dimensions[column_letter].width = min(max_width, max_length)

    # Iterate over data elements
    row = 2
    for item in data['items']:
        question = item['question']['text']
        chunks = item['chunks']['items']
        answers = item['answers']['items']

        # Write the question
        worksheet.cell(row=row, column=1, value=question)

        # Write chunks into separate columns
        for col, chunk in enumerate(chunks, start=2):
            try:
                worksheet.cell(row=row, column=col, value=chunk['text'])
            except openpyxl.utils.exceptions.IllegalCharacterError:
                cleaned_text = ILLEGAL_CHARACTERS_RE.sub('', chunk['text'])
                worksheet.cell(row=row, column=col, value=cleaned_text)

        # Write LLM responses and human evaluations
        for answer_idx, answer in enumerate(answers, start=1):
            answer_text = answer['text'] # Extraire le texte de la réponse à partir de 'text'
            worksheet.cell(row=row, column=max_chunks+2, value=f"Answer {answer_idx}: {answer_text}")
            worksheet.cell(row=row, column=max_chunks+3, value=answer['eval']['human'])
            row += 1

    # Adjust column widths
    set_column_width(worksheet, 1, max_length=50)   # For questions
    for col in range(2, column):
        set_column_width(worksheet, col)  # For chunks
    set_column_width(worksheet, column, max_length=100)  # For LLM responses
    set_column_width(worksheet, column + 1, max_length=100)  # For human evaluations

    # Save the Excel workbook
    workbook.save(xlsx_path)


def xlsx_to_json(json_path : str, xlsx_path : str):
    # Open the XLSX file
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active

    # Open the JSON file
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Find columns for LLM responses and human evaluations
    answer_column = None
    eval_column = None
    for col in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=col).value
        if value == 'LLM answers':
            answer_column = col
        elif value == 'Human evaluations':
            eval_column = col

    # Find the first row containing a question
    start_row = None
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value:
            start_row = row
            break

    # Update human evaluations in JSON data
    row = start_row
    question_index = 0
    while row <= worksheet.max_row:
        item = data['items'][question_index]
        answers = item['answers']['items']

        for answer_idx in range(len(answers)):
            answer_cell = worksheet.cell(row=row, column=answer_column)
            eval_cell = worksheet.cell(row=row, column=eval_column)

            if answer_cell.value and eval_cell.value:
                answer_text = answers[answer_idx]['text'] 
                answers[answer_idx]['eval']['human'] = eval_cell.value

            row += 1

        question_index += 1
    
    # Extract the directory path and filename from the JSON path
    json_dirname = os.path.dirname(json_path)
    json_filename = os.path.basename(json_path)
    # Remove the ".json" extension
    json_filename_updated = os.path.splitext(json_filename)[0] + "updated.json"
    # Construct the full path for the XLSX file in the same directory
    json_path = os.path.join(json_dirname, json_filename_updated)    # Remove the ".json" extension
    # Save the updated JSON data
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)